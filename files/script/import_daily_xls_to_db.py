#!/usr/bin/env python3
"""Import TheBook's exported .xlsx file into its SQLite database.

This version uses openpyxl and therefore requires a real Office Open XML
workbook. The Android exporter uses JXL and writes legacy BIFF .xls content;
historical exports may still have an .xlsx filename. Android-generated backups
must be converted to OOXML before they can be read by this script.

Install the only third-party dependency with:

    python -m pip install openpyxl

Examples:

    # Append records, skipping rows already present by title + timestamp.
    python import_daily_xls_to_db.py export.xlsx --db daily.db

    # Replace all daily/stat_month data with the contents of the export.
    python import_daily_xls_to_db.py export.xlsx --db daily.db --mode replace
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Iterable


DAILY_COLUMNS = ("title", "year", "month", "day", "hour", "time")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import TheBook exported .xlsx daily records into daily.db."
    )
    parser.add_argument("xls_file", type=Path, help="The exported .xlsx file.")
    parser.add_argument(
        "--db",
        type=Path,
        default=Path("daily.db"),
        help="SQLite database path; default: daily.db",
    )
    parser.add_argument(
        "--mode",
        choices=("append", "replace"),
        default="append",
        help="append records or replace daily/stat_month; default: append",
    )
    return parser.parse_args()


def as_int(value: object, field: str, row_number: int) -> int:
    """Convert spreadsheet cell values to integers with a useful error message."""
    if value is None or str(value).strip() == "":
        raise ValueError(f"row {row_number}: {field} is empty")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"row {row_number}: {field} must be an integer, got {value!r}"
        ) from exc


def load_xlsx(path: Path) -> list[tuple[str, int, int, int, int, int]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency openpyxl. Install it with: "
            "python -m pip install openpyxl"
        ) from exc

    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Cannot read {path} as a real .xlsx file. "
            "The Android JXL export is legacy .xls content."
        ) from exc

    try:
        sheet = workbook["daily"] if "daily" in workbook.sheetnames else workbook.worksheets[0]
        records: list[tuple[str, int, int, int, int, int]] = []

        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            values = list(row)
            if not values or all(str(value).strip() == "" for value in values):
                continue
            if len(values) < 7:
                raise ValueError(
                    f"row {row_index + 1}: expected at least 7 columns "
                    "(id, title, year, month, day, hour, time)"
                )

            # The Android exporter writes no header. Accepting a header makes
            # the script tolerant of manually edited exports.
            if row_index == 0 and str(values[0]).strip().lower() == "id":
                continue

            title = str(values[1]).strip()
            if not title:
                raise ValueError(f"row {row_index + 1}: title is empty")

            year = as_int(values[2], "year", row_index + 1)
            month = as_int(values[3], "month", row_index + 1)
            day = as_int(values[4], "day", row_index + 1)
            hour = as_int(values[5], "hour", row_index + 1)
            timestamp = as_int(values[6], "time", row_index + 1)

            if not 1 <= month <= 12:
                raise ValueError(f"row {row_index + 1}: month out of range: {month}")
            if not 1 <= day <= 31:
                raise ValueError(f"row {row_index + 1}: day out of range: {day}")
            if not 0 <= hour <= 23:
                raise ValueError(f"row {row_index + 1}: hour out of range: {hour}")

            records.append((title, year, month, day, hour, timestamp))

        return records
    finally:
        workbook.close()


def ensure_schema(connection: sqlite3.Connection) -> None:
    # These definitions match the Room entities Daily and Stat.
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS daily (
            title TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            day INTEGER NOT NULL,
            hour INTEGER NOT NULL,
            time INTEGER NOT NULL,
            id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS stat_month (
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            times INTEGER NOT NULL,
            tag TEXT NOT NULL,
            id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL
        );
        """
    )


def append_records(
    connection: sqlite3.Connection,
    records: Iterable[tuple[str, int, int, int, int, int]],
) -> tuple[int, int]:
    inserted = 0
    skipped = 0
    for record in records:
        title, year, month, day, hour, timestamp = record
        exists = connection.execute(
            """
            SELECT 1 FROM daily
            WHERE title = ? AND year = ? AND month = ? AND day = ?
              AND hour = ? AND time = ?
            LIMIT 1
            """,
            record,
        ).fetchone()
        if exists:
            skipped += 1
            continue
        connection.execute(
            """
            INSERT INTO daily(title, year, month, day, hour, time)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (title, year, month, day, hour, timestamp),
        )
        inserted += 1
    return inserted, skipped


def rebuild_month_stats(connection: sqlite3.Connection) -> None:
    """Rebuild derived monthly statistics from the source daily table."""
    counts = connection.execute(
        """
        SELECT year, month, COUNT(*)
        FROM daily
        GROUP BY year, month
        ORDER BY year, month
        """
    ).fetchall()
    connection.execute("DELETE FROM stat_month")
    connection.executemany(
        "INSERT INTO stat_month(year, month, times, tag) VALUES (?, ?, ?, ?)",
        (
            (year, month, count, f"{year:04d}{month:02d}")
            for year, month, count in counts
        ),
    )


def main() -> int:
    args = parse_args()
    try:
        records = load_xlsx(args.xls_file)
        args.db.parent.mkdir(parents=True, exist_ok=True)

        with sqlite3.connect(args.db) as connection:
            ensure_schema(connection)
            if args.mode == "replace":
                connection.execute("DELETE FROM daily")
                connection.execute("DELETE FROM stat_month")
                inserted, skipped = append_records(connection, records)
            else:
                inserted, skipped = append_records(connection, records)
            rebuild_month_stats(connection)

            daily_count = connection.execute("SELECT COUNT(*) FROM daily").fetchone()[0]
            month_count = connection.execute("SELECT COUNT(*) FROM stat_month").fetchone()[0]

        print(f"Imported file : {args.xls_file}")
        print(f"Database      : {args.db}")
        print(f"Mode          : {args.mode}")
        print(f"Rows read     : {len(records)}")
        print(f"Rows inserted : {inserted}")
        print(f"Rows skipped  : {skipped}")
        print(f"Daily total   : {daily_count}")
        print(f"Month stats   : {month_count}")
        return 0
    except (FileNotFoundError, RuntimeError, ValueError, sqlite3.Error) as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
